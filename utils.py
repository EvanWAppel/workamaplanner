import os
from datetime import datetime, date, timedelta
import pandas as pd
import re
import calendar
import os.path
import requests
import utils
import json
from enum import Enum
import dateutil.relativedelta as rd
import time
import numpy as np
from openpyxl import load_workbook

def logger():
    if os.environ['OS'] == "Windows_NT":
        log = r"c:\Users\{}\efficiency\log.txt".format(os.environ['USERNAME'])
    elif os.environ['OS'] == 'POSIX':
        log = r"/homelog.txt".format(os.environ['HOME'])
    else:
        print(f"{datetime.today()} >>> Here's a nickel, kid. Go and buy yourself a better computer.")
    return log

def entry(text):
    log = logger()
    message = f">>> {datetime.today()} ]|[ {text}"
    with open(log,'a') as l:
        print(message,file=l)
    print(message,flush=True)
    l.close()
class Mode(Enum):
    marcom = {"nt":
                    {"destination":r"",
                     "template_name":r"marcom_template.xlsx"}}
class assumptions():
    def __init__(self,mode:utils.Mode = utils.Mode.marcom):
        self.year = date.today().year
        self.last = calendar.monthrange(date.today().year, date.today().month)[1]
        self.month = date.today().month
        if date.today().month == 12:
            self.month_two = 0
            self.year_two = date.today().year + 1
        else:
            self.month_two = date.today().month
            self.year_two = date.today().year
        if date.today().day <= 14:
            self.default_start = datetime(self.year,self.month,15)
            self.default_end = datetime(self.year,self.month,self.last)
        else:
            self.default_start = datetime(self.year_two,self.month_two+1,1)
            self.default_end = datetime(self.year_two,self.month_two+1,14)
        if mode.name == "marcom":
            self.dest_dir = mode.value[os.name]["destination"]
            self.template_name = mode.value[os.name]["template_name"]


        self.current_directory = os.getcwd()
        self.headers: dict = {'APIAccessToken':r""
                            ,'UserToken':r''}
        self.timestamp = datetime.now()
        self.history_in_days = 365
        
    class apis(Enum):
        cust_report_str = r"https://app6.workamajig.com/api/beta1/reports?ReportKey=Sy9DZTVraGlUMTU2clVOYVY0cnhsZz090"
        emp_api = r"https://app6.workamajig.com/api/beta1/employees"
        time_api = r"https://app6.workamajig.com/api/beta1/time?userKey={}&includeTime=1"
        proj_api = r"https://app6.workamajig.com/api/beta1/projects?id={}&includeTasks=true"
        task_api = r"https://app6.workamajig.com/api/beta1/tasks?projectNumber={}"
    
def get_call(assumptions: utils.assumptions,api_address: str) -> tuple:
    payload = requests.get(api_address,headers=assumptions.headers)
    utils.entry(f"Status code: {payload.status_code} {payload.reason}")
    content = json.loads(payload.content)
    return (payload, content)

def employee_list(payload: dict) -> list:
    employees = []
    for i in payload['data']['employee']:
        employees.append(i['employeeKey'])
    utils.entry(f"Employee list length: {len(employees)}")
    return employees

def timesheet_retrieval(employees: list, data: utils.assumptions) -> dict:
    sheets = pd.DataFrame()
    times = []
    for j in employees:
        response = utils.get_call(assumptions = data, api_address = data.apis.time_api.value.format(j))
        utils.entry(f"Reason: {response[0].reason}, Response: {response[0].status_code}, Elapsed time: {response[0].elapsed}, Employee ID: {j}")
        try:
            for k in response[1]['data']['timesheet']:
                sheets = pd.concat([sheets, pd.DataFrame(k)])
                try:
                    for l in k['TimeEntries']:
                        times.append(l)
                except:
                    entry(f"No Time Entries.")
        except:
            entry(f"Employee {j} has no timesheets.")
    time_entries = pd.DataFrame(times )
    return {"raw_response":response,'sheets':sheets,'time_entries':time_entries}

# def time_entry_extractor(sheets: pd.DataFrame) -> pd.DataFrame:
#     times = pd.DataFrame()

#     for k in sheets['TimeEntries']:
#         times = pd.concat([times,pd.DataFrame(k,index=[0])])
    
#     return times

def pre_tasker(sheets: pd.DataFrame, times: pd.DataFrame) -> pd.DataFrame:
    # The username isn't on the time entry JSON, so we rejoin the data from sheets to name them.
    names = sheets[['TimeSheetKey','UserName']].drop_duplicates()
    # Base table is the timesheets and time sessions of all employees since the inception of Workamajig
    base = pd.merge(times,names,how="left",left_on="timeSheetKey",right_on="TimeSheetKey")
    utils.entry(f"Cleaning up the table a little. {len(base)} rows.")
    return base

def tasker(base: pd.DataFrame) -> pd.DataFrame:
    # Convert the time entries into tasks so that you can figure out the estimates.
    tasks = base.groupby(["projectFullName",
                        "projectNumber",
                        "taskID",
                        "workDate",
                        "UserName",
                        "serviceCode",
                        "serviceDescription",
                        "taskName", 
                        "campaignName",
                        "clientName",
                        "taskKey",
                        ],as_index=False).agg({"actualHours":"sum"})
    utils.entry(f"Tasks Assembled. {len(tasks)}")
    return tasks

def estimator(tasks: pd.DataFrame) -> pd.DataFrame:
    # Create Estimates
    estimates = tasks.groupby(['UserName','taskName','clientName'],as_index=False).agg({'actualHours':"mean"})
    estimates = estimates.rename(columns={"actualHours":"estimate"})
    utils.entry(f"Estimate table created. {len(estimates)} rows.")
    return estimates

def api_cust():
    # makes the datetime a record in the operating system environment variables
    os.environ['API_attempt_at_'+str(datetime.now())] = "1"
    ls = []
    # Looks through the variables for...
    for i in dict(os.environ).keys():
        # our date records...
        if i.__contains__('API_ATTEMPT'):
            # and turns them back into dates But not before making a list of variables that are just sitting out there)
            ls.append(i)
            # ...Checks to see if the date of the env variable record is older than 15 mins and then drops it if it is.
            if datetime.strptime(i[15:],"%Y-%m-%d %H:%M:%S.%f") < (datetime.now() - rd.relativedelta(minutes = 15)):
                del os.environ[i]
    # Depending on how many env variables exist there are so many tries left.
    if len(ls) == 0:
        m = "Five calls remaining."
    elif len(ls) == 1:
        m = "Four calls remaining."
    elif len(ls) == 2:
        m = "Three calls remaining."
    elif len(ls) == 3:
        m = "Two calls remaining."
    elif len(ls) == 4:
        m = "One call remaining."
    else:
        # prints some information about when you can try again.
        m = "Hey, you've overloaded the API. Ease off until "+ str(datetime.strptime(min(ls)[26:],"%H:%M:%S.%f") + rd.relativedelta(minutes = 15))[12:]
    utils.entry(f"{m}")
    return m

# Create list of all projects
def cust_report(data: utils.assumptions) -> dict:
    api_cust()
    response = utils.get_call(assumptions = data, api_address = utils.assumptions.apis.cust_report_str.value)
    project_table = pd.DataFrame(response[1]['data']['report'])
    return {"response":response,"project_table":project_table}

def proj_numbers(project_table: pd.DataFrame, data: utils.assumptions) -> list:
    # Date needs to be in Datetsime format
    project_table['first_Task_Start_Date'] = pd.to_datetime(project_table['first_Task_Start_Date'])
    # Project list of just the last year
    proj_list = project_table[(project_table['first_Task_Start_Date']>(data.timestamp - timedelta(days=data.history_in_days)))]
    # Only project number
    proj_list = proj_list[['project_Number']]
    # Return a copy
    proj_list = proj_list['project_Number'].tolist()
    return proj_list

# Collect tasks for all projects
def get_tasks(projects: list, data: utils.assumptions, sleeper: float = 2.8) -> tuple:
    x: int = 1
    payloads = {}
    task_collection: pd.DataFrame = pd.DataFrame()
    for i in projects:
        entry(f"Now collecting data for project {i}, {x} of {len(projects)}...")
        x = x + 1 
        response = utils.get_call(assumptions = data, api_address = utils.assumptions.apis.proj_api.value.format(i))
        contentdatatasks = response[1]['data']['tasks']
        frameddatatasks = pd.DataFrame(contentdatatasks)
        complete_data = pd.DataFrame(utils.get_call(assumptions = data, api_address = utils.assumptions.apis.task_api.value.format(i))[1]['data']['task'])[["taskKey","actComplete"]]
        tasks = pd.merge(frameddatatasks,complete_data,how='left',on='taskKey')
        tasks = tasks.assign(project_Number = i)
        userlist: list = []
        for j in contentdatatasks:
            user = j['taskUser']
            for jj in user:
                userlist.append(jj)
        if userlist != 0:
            try:
                tasks = pd.merge(tasks, pd.DataFrame(userlist), how='left', on="taskKey")
            except Exception as e:
                entry(f"{e}")
        task_collection = pd.concat([task_collection,tasks])
        payloads.update({"payload_for_"+str(i):response[0]})
        time.sleep(sleeper)
    return (task_collection, payloads)

def estimate_and_task_join(project_list: list, taskset: pd.DataFrame, estimates: pd.DataFrame, data: utils.assumptions) -> dict:
    # Apply estimates to full list of tasks
    pl = pd.DataFrame(project_list,columns=["projectNumber"])
    pl['projectNumber'] = pl['projectNumber'].astype('str')
    taskset['projectNumber'] = taskset['projectNumber'].astype('str').str.replace(".0","")
    master = pd.merge(pl,taskset,how='left',on='projectNumber')
    estimates = estimates.rename(columns={'UserName':"userName"})
    estimates = estimates.groupby(['userName','taskName'],as_index=False).agg({'estimate':np.mean})
    master = pd.merge(master, estimates,how='left',on=["userName","taskName"])
    master['actStart'] = pd.to_datetime(master['actStart']).dt.tz_localize(None)
    master['actComplete'] = pd.to_datetime(master['actComplete']).dt.tz_localize(None)
    master['planStart'] = pd.to_datetime(master['planStart']).dt.tz_localize(None)
    master['planComplete'] = pd.to_datetime(master['planComplete']).dt.tz_localize(None)
    master = master[master['taskName'].str.contains('[Cc}]lient')!=True]
    unassigned = master[master['userName'].isna()==True]
    unassigned = master[(master['planStart']>=data.default_start)&(master['planStart']>=data.default_end)]
    unassigned = unassigned[['projectNumber']].drop_duplicates()
    return {"master":master,"unassigned":unassigned}

# Print full dataset as DATA to the capacity planner
def summary_table(dataset: dict,data: utils.assumptions) -> str:
    # Then, overwrite the data tab.
    def write_excel(filename,sheetname,dataframe):
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
            try:
                writer.book.remove(writer.book[sheetname])
            except:
                print("Worksheet does not exist")
            finally:
                dataframe.to_excel(writer, sheet_name=sheetname,index=False)
                #writer.save()
    write_excel(os.path.join(data.dest_dir,data.template_name),"DATA",dataset['master'])
    write_excel(os.path.join(data.dest_dir,data.template_name),"UNASSIGNED",dataset['unassigned'])
    # Write variables into the spreadsheet
    wb = load_workbook(os.path.join(data.dest_dir,data.template_name),read_only=False)
    ws = wb["VARIABLES"]
    ws["A1"] = str("Variable")
    ws["A2"] = str("Start")
    ws["B2"] = str(data.default_start)
    ws["A3"] = str("End")
    ws["B3"] = str(data.default_end)
    ws["A4"] = str("Period Hours")
    if data.default_start.strftime("%d") =='01':
        b = "Early {} {}".format(data.default_start.strftime("%B"),data.default_start.strftime("%Y"))
    elif data.default_start.strftime("%d") == '15':
        b = "Late {} {}".format(data.default_start.strftime("%B"),data.default_start.strftime("%Y"))
    else:
        b = "On {}".format(data.default_start.strftime("%d%B%Y"))
    wb.save(os.path.join(data.dest_dir,r"planners\planner {}.xlsx".format(b)))
    wb.save(os.path.join(data.dest_dir,data.template_name))
    fin = "CAPLAN has completed. Have a nice day."
    entry(f"{fin}")
    return f"{fin}"